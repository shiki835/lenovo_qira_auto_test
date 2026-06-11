"""Test executor — runs questions sequentially against the chat client.

Key features:
- Sequential execution with progress display
- Incremental saves to JSON (resume after interruption)
- Automatic error recovery (retry, restart app, timeout reset)
- Time estimation for long runs
- Writes elapsed time back to Excel
"""

import json
import time
import logging
from pathlib import Path
from datetime import datetime, timezone
from dataclasses import dataclass, field, asdict

from .config import Config
from .excel_reader import QuestionSet, load_questions
from .ui_automator import UIAutomator

logger = logging.getLogger(__name__)


@dataclass
class ResultEntry:
    question: str
    expected: str
    category: str
    response: str
    screenshot_path: str
    elapsed_seconds: float
    error: str
    retries: int


@dataclass
class RunResults:
    version: str
    source_file: str
    started_at: str
    completed_at: str = ""
    total: int = 0
    completed: int = 0
    results: dict[int, ResultEntry] = field(default_factory=dict)


class Executor:
    """Runs test questions against the chat client."""

    def __init__(self, config: Config, version: str):
        self.config = config
        self.version = version
        self.results_dir = Path(__file__).parent.parent / "results" / version
        self.results_path = self.results_dir / "raw_responses.json"
        self.automator: UIAutomator | None = None

    def run(self, excel_path: str) -> RunResults:
        """Execute all questions from the Excel file."""
        questions = load_questions(
            excel_path,
            question_column=self.config.excel.question_column,
            expected_column=self.config.excel.expected_column,
            category_column=self.config.excel.category_column,
            reset_session_column=self.config.excel.reset_session_column,
        )

        results = self._load_existing_results() or RunResults(
            version=self.version,
            source_file=excel_path,
            started_at=datetime.now(timezone.utc).isoformat(),
            total=len(questions),
        )

        self.automator = UIAutomator(self.config, self.results_dir)

        pending = [q for q in questions if q.index not in results.results]
        if not pending:
            logger.info("All questions already completed — nothing to run.")
            return results

        logger.info(f"Starting run: {len(questions)} total, {len(pending)} pending")

        if not self.automator.ensure_app_ready():
            raise RuntimeError(
                "Could not find or launch the chat client."
            )

        # Start fresh only when the next pending question starts a new session.
        if pending[0].reset_session and self.automator.has_reset():
            print("Resetting to start fresh...")
            self.automator.reset_chat()

        results.total = len(questions)
        start_time = time.time()

        for i, q in enumerate(pending):
            elapsed_total = time.time() - start_time
            completed = results.completed
            eta = self._estimate_eta(elapsed_total, completed, results.total)

            print(f"\r[{completed}/{results.total}] "
                  f"Elapsed: {self._fmt_time(elapsed_total)} "
                  f"ETA: {eta} "
                  f"Next: {q.question[:40]}...", end="", flush=True)

            entry = self._run_one(q)
            results.results[q.index] = entry
            results.completed += 1
            self._save_results(results)

            # Recover UI after an exhausted timeout without destroying non-new context.
            if entry.error in ("TIMEOUT", "SEND_CLICK_FAILED"):
                print(f"\n  Q{q.index} {entry.error}, recovering chat...")
                try:
                    self.automator.recover_after_failure(q.reset_session)
                except Exception:
                    pass

        # End: reset to clean up
        print("\nFinal reset...")
        try:
            self.automator.reset_chat()
        except Exception:
            pass

        results.completed_at = datetime.now(timezone.utc).isoformat()
        self._save_results(results)
        self._write_elapsed_to_excel(excel_path, results)

        total_elapsed = time.time() - start_time
        print(f"\n\nDone! {results.completed}/{results.total} completed "
              f"in {self._fmt_time(total_elapsed)}")
        self._print_summary(results)

        return results

    def resume(self, excel_path: str) -> RunResults:
        return self.run(excel_path)

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _run_one(self, question) -> ResultEntry:
        max_retries = self.config.execution.retry_count + 1
        last_error = ""

        for attempt in range(max_retries):
            try:
                resp = self.automator.ask_question(
                    question.question, question.index,
                    reset_first=question.reset_session,
                )
                if resp.error:
                    last_error = resp.error
                    if resp.error in ("TIMEOUT", "SEND_BUTTON_NOT_FOUND", "SEND_CLICK_FAILED"):
                        logger.warning(
                            f"Q{question.index}: {resp.error}, "
                            f"attempt {attempt + 1}/{max_retries}"
                        )
                        if attempt < max_retries - 1:
                            try:
                                self.automator.recover_after_failure(question.reset_session)
                            except Exception:
                                logger.exception("Recovery after failure failed")
                            continue
                        return ResultEntry(
                            question=question.question,
                            expected=question.expected,
                            category=question.category,
                            response="",
                            screenshot_path=resp.screenshot_path,
                            elapsed_seconds=resp.elapsed_seconds,
                            error=resp.error,
                            retries=attempt,
                        )
                    if resp.error == "EMPTY_RESPONSE":
                        logger.warning(
                            f"Q{question.index}: empty response, "
                            f"attempt {attempt + 1}/{max_retries}"
                        )
                        self.automator.restart_app()
                        self.automator.ensure_app_ready()
                        continue
                    return ResultEntry(
                        question=question.question,
                        expected=question.expected,
                        category=question.category,
                        response=resp.text,
                        screenshot_path=resp.screenshot_path,
                        elapsed_seconds=resp.elapsed_seconds,
                        error=resp.error,
                        retries=attempt,
                    )
                return ResultEntry(
                    question=question.question,
                    expected=question.expected,
                    category=question.category,
                    response=resp.text,
                    screenshot_path=resp.screenshot_path,
                    elapsed_seconds=resp.elapsed_seconds,
                    error="",
                    retries=attempt,
                )
            except Exception as e:
                last_error = str(e)
                logger.error(f"Q{question.index} attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    try:
                        self.automator.recover_after_failure(question.reset_session)
                    except Exception:
                        logger.exception("Recovery after exception failed")

        return ResultEntry(
            question=question.question,
            expected=question.expected,
            category=question.category,
            response="",
            screenshot_path="",
            elapsed_seconds=0,
            error=f"ALL_RETRIES_EXHAUSTED: {last_error}",
            retries=max_retries - 1,
        )

    def _write_elapsed_to_excel(self, excel_path: str, results: RunResults):
        """Save a copy of the Excel with elapsed time and errors in the results dir."""
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        elapsed_col = self._ensure_column(ws, "回复耗时(s)")
        error_col = self._ensure_column(ws, "错误信息")

        for row in ws.iter_rows(min_row=2):
            idx = row[0].row - 1
            if idx in results.results:
                entry = results.results[idx]
                ws.cell(row=row[0].row, column=elapsed_col,
                        value=round(entry.elapsed_seconds, 1))
                ws.cell(row=row[0].row, column=error_col,
                        value=entry.error if entry.error else "")

        out_path = self.results_dir / "test_results.xlsx"
        wb.save(str(out_path))
        wb.close()

    @staticmethod
    def _ensure_column(ws, col_name):
        """Add column header if missing, return its 1-based index."""
        for cell in ws[1]:
            if cell.value and str(cell.value).strip() == col_name:
                return cell.column
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=col_name)
        return col

    def _load_existing_results(self) -> RunResults | None:
        if not self.results_path.exists():
            return None
        try:
            with open(self.results_path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            results = RunResults(
                version=raw["version"],
                source_file=raw["source_file"],
                started_at=raw["started_at"],
                completed_at=raw.get("completed_at", ""),
                total=raw["total"],
                completed=raw["completed"],
            )
            for k, v in raw.get("results", {}).items():
                results.results[int(k)] = ResultEntry(**v)
            return results
        except Exception as e:
            logger.warning(f"Could not load existing results: {e}, starting fresh")
            return None

    def _save_results(self, results: RunResults):
        self.results_dir.mkdir(parents=True, exist_ok=True)
        payload = {
            "version": results.version,
            "source_file": results.source_file,
            "started_at": results.started_at,
            "completed_at": results.completed_at,
            "total": results.total,
            "completed": results.completed,
            "results": {
                str(k): asdict(v) for k, v in results.results.items()
            },
        }
        tmp = str(self.results_path) + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        Path(tmp).replace(self.results_path)

    def _print_summary(self, results: RunResults):
        errors = sum(1 for r in results.results.values() if r.error)
        timeouts = sum(1 for r in results.results.values() if r.error == "TIMEOUT")
        print(f"  Success: {results.completed - errors}")
        print(f"  Errors:  {errors} (timeouts: {timeouts})")
        avg_elapsed = sum(
            r.elapsed_seconds for r in results.results.values()
        ) / max(results.completed, 1)
        print(f"  Avg response time: {avg_elapsed:.1f}s")

    @staticmethod
    def _estimate_eta(elapsed: float, completed: int, total: int) -> str:
        if completed == 0:
            return "calculating..."
        avg = elapsed / completed
        remaining = (total - completed) * avg
        return Executor._fmt_time(remaining)

    @staticmethod
    def _fmt_time(seconds: float) -> str:
        m, s = divmod(int(seconds), 60)
        if m > 0:
            return f"{m}m{s}s"
        return f"{s}s"
